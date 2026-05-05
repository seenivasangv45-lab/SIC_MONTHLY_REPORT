"""
Combined Excel Report Processor
================================
Template lives in  ./template/  folder on disk (no upload needed).
User uploads multiple raw data files — each matched by filename prefix:
  e.g.  ARC_18_abc.xlsx           → ARC processor
        PCD_20_xyz.xlsx           → PCD processor
        CNT_27_BySvcDate.xlsx     → template filler (CNT_27)
        CNT_19_....xlsx           → template filler (CNT_19)
        PAT_2_BySvcDate_41c9.xlsx → template filler (PAT_2)
        FIN_18_....xlsx           → template filler (FIN_18)

Output: single .xlsx with up to 4 sheets:
  1. Begin & Ending AR by Clinic
  2. SJC_Closed Report
  3. SJC_E&M_Visit Summary
  4. Charge Detail Report  (requires template + CNT_27/19, PAT_2, PCD_20, FIN_18)


Usage:  python app.py
"""

import os, re, uuid, io, glob
import pandas as pd
from openpyxl import load_workbook
from flask import Flask, request, jsonify, send_file, render_template

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200 MB total

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FOLDER = os.path.join(BASE_DIR, "outputs")
TEMPLATE_DIR  = os.path.join(BASE_DIR, "template")   # drop your template.xlsx here

os.makedirs(TEMPLATE_DIR, exist_ok=True)

# ──────────────────────────────────────────────────────────────────────────────
#  File resolver — match uploaded filename → data-source key by prefix
#  Rule: strip extension, uppercase, then check if it equals a prefix OR
#        starts with prefix followed by a non-alphanumeric character.
#  Examples:
#    PCD_20.xlsx                → "PCD_20"
#    PCD_20_BySvcDate.xlsx      → "PCD_20"
#    PAT_2_BySvcDate_41c9.xlsx  → "PAT_2"
#    CNT_271_extra.xlsx         → None  (271 ≠ 27_...)
# ──────────────────────────────────────────────────────────────────────────────

# Longer/more-specific prefixes listed first
RAW_PREFIXES = ["ARC_18", "PCD_20", "CNT_27", "CNT_19", "PAT_2", "FIN_18"]

def resolve_file_key(filename: str) -> str | None:
    stem = os.path.splitext(filename)[0].upper()   # e.g. "PAT_2_BYSVCDATE_41C9BDED"
    for prefix in RAW_PREFIXES:
        p = prefix.upper()
        if stem == p:
            return prefix
        # prefix must be followed by a separator, not another alphanumeric char
        if stem.startswith(p) and len(stem) > len(p) and not stem[len(p)].isalnum():
            return prefix
    return None


def find_template() -> str | None:
    """Return path of the first .xlsx/.xls found in ./template/, or None."""
    for ext in ("*.xlsx", "*.xls"):
        matches = glob.glob(os.path.join(TEMPLATE_DIR, ext))
        if matches:
            return matches[0]
    return None


# ──────────────────────────────────────────────────────────────────────────────
#  Number format strings
# ──────────────────────────────────────────────────────────────────────────────
FMT_DOLLAR  = '$#,##0.00_);($#,##0.00)'
FMT_INTEGER = '#,##0'
FMT_DATE    = 'mm/dd/yyyy'
FMT_TEXT    = '@'

# ──────────────────────────────────────────────────────────────────────────────
#  Shared reader — always reads the first sheet of a raw file
# ──────────────────────────────────────────────────────────────────────────────

def read_first_sheet(file_bytes: bytes) -> pd.DataFrame:
    try:
        return pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, engine="calamine")
    except Exception:
        return pd.read_excel(io.BytesIO(file_bytes), sheet_name=0)


# ──────────────────────────────────────────────────────────────────────────────
#  ARC_18 processing
# ──────────────────────────────────────────────────────────────────────────────

def parse_currency(val):
    if pd.isna(val): return 0.0
    s = str(val).replace("$","").replace(",","").strip()
    if s.startswith("(") and s.endswith(")"): s = "-" + s[1:-1]
    try:    return float(s)
    except: return 0.0

CLINIC_TO_LOCATION = {
    "BLACKCREEK":"BLACKCREEK","BLUFFTON":"BLUFFTON","HEARTWOOD":"HEARTWOOD",
    "HINESVILLE":"HINESVILLE","POOLER95":"POOLER95","POOLER16":"POOLER16",
    "RINCON":"RINCON","RICHMOND":"RICHMOND","SAVANNAH":"SAVANNAH","STATESBORO":"STATESBORO",
}

def normalise_clinic(v):
    key = re.sub(r"[^A-Z]","",str(v).upper())
    for l in range(len(key),0,-1):
        if key[:l] in CLINIC_TO_LOCATION: return CLINIC_TO_LOCATION[key[:l]]
    return str(v).strip().upper()

EPS_CLASS      = "9-EPS"
SELF_PAY_CLASS = "1-Self Pay"
NUMERIC_COLS   = ["textbox1","textbox18","textbox19","textbox214","textbox20","textbox21"]

def get_previous_month_range():
    """Return (start, end) of the previous calendar month relative to today."""
    today = pd.Timestamp.today().normalize()
    first_of_this_month = today.replace(day=1)
    last_of_prev = first_of_this_month - pd.Timedelta(days=1)
    first_of_prev = last_of_prev.replace(day=1)
    return first_of_prev, last_of_prev


def compute_patient_counts_from_cnt27(cnt27_bytes: bytes) -> dict:
    """Return {normalised_clinic: unique Pat_Num count} filtered to previous month."""
    df = read_first_sheet(cnt27_bytes)
    df["Svc_Date"] = pd.to_datetime(df["Svc_Date"], errors="coerce")
    start, end = get_previous_month_range()
    mask = (df["Svc_Date"] >= start) & (df["Svc_Date"] <= end)
    df = df[mask].copy()
    df["_clinic_key"] = df["Clinic"].apply(normalise_clinic)
    return (
        df.groupby("_clinic_key")["Pat_Num"]
          .nunique()
          .to_dict()
    )


def compute_arc_summary(file_bytes: bytes, cnt27_bytes: bytes = None) -> pd.DataFrame:
    patient_counts = {}
    if cnt27_bytes is not None:
        try:
            patient_counts = compute_patient_counts_from_cnt27(cnt27_bytes)
        except Exception:
            pass
    df = read_first_sheet(file_bytes)
    for col in NUMERIC_COLS:
        df[col+"_n"] = df[col].apply(parse_currency)
    df["Location"] = df["Clinic1"].apply(normalise_clinic)
    rows = []
    for loc in sorted(df["Location"].unique()):
        ld  = df[df["Location"]==loc]
        ep  = ld[ld["Financial_Class"]==EPS_CLASS]
        sp  = ld[ld["Financial_Class"]==SELF_PAY_CLASS]
        pay = ld[~ld["Financial_Class"].isin([SELF_PAY_CLASS,EPS_CLASS])]
        ba  = ld["textbox1_n"].sum();  ea = ld["textbox21_n"].sum()
        co  = ld["textbox19_n"].sum(); ot = ld["textbox214_n"].sum()
        rows.append({
            "Location":ld["Location"].iloc[0],"Dpt Name":None,"Dpt #":None,
            "Biginning AR":ba,"Total Charges":ld["textbox18_n"].sum(),
            "Payer Charges":pay["textbox18_n"].sum(),"Self-pay charges":sp["textbox18_n"].sum(),
            "EPS":ep["textbox18_n"].sum(),"Total Payments":ld["textbox20_n"].sum(),
            "Payer Payments":pay["textbox20_n"].sum(),"Patient Payments":sp["textbox20_n"].sum(),
            "EPS.1":ep["textbox20_n"].sum(),"Contractual adjustment":co,
            "Other Adjustment":ot,"Total contractual":co+ot,"Refund":0.00,
            "Patient count":patient_counts.get(ld["Location"].iloc[0], 0),
            "claim count":int(ld["Primary_Visit_Count"].sum()),
            "Ending AR":ea,"Change in AR":ba-ea,
        })
    s = pd.DataFrame(rows)
    num_sum = ["Biginning AR","Total Charges","Payer Charges","Self-pay charges","EPS",
               "Total Payments","Payer Payments","Patient Payments","EPS.1",
               "Contractual adjustment","Other Adjustment","Total contractual",
               "Refund","Patient count","claim count","Ending AR","Change in AR"]
    total = {c:(s[c].sum() if c in num_sum else ("Total" if c=="Location" else None)) for c in s.columns}
    total["claim count"] = int(total["claim count"])
    total["Patient count"] = int(total["Patient count"])
    total["Refund"] = float(total["Refund"])
    return pd.concat([s, pd.DataFrame([total])], ignore_index=True)

ARC_COL_SPEC = {
    "Location":(None,18),"Dpt Name":(None,14),"Dpt #":(None,8),
    "Biginning AR":("dollar",16),"Total Charges":("dollar",16),
    "Payer Charges":("dollar",16),"Self-pay charges":("dollar",18),
    "EPS":("dollar",14),"Total Payments":("dollar",16),
    "Payer Payments":("dollar",16),"Patient Payments":("dollar",16),
    "EPS.1":("dollar",14),"Contractual adjustment":("dollar",22),
    "Other Adjustment":("dollar",18),"Total contractual":("dollar",18),
    "Refund":("dollar",12),"Patient count":("integer",14),
    "claim count":("integer",14),"Ending AR":("dollar",16),"Change in AR":("dollar",16),
}

# ──────────────────────────────────────────────────────────────────────────────
#  PCD_20 processing
# ──────────────────────────────────────────────────────────────────────────────

OUTPUT1_COLS = ["Inv_Num","Svc_Date","Proc_Code","Modifier","Pat_Name",
                "Pat_Num","Pat_DOB","textbox13","textbox11","textbox9",
                "Phy_Name","Crg_Amt","Adj_Amt","Paid_Amt","Balance"]
OUTPUT2_FINAL_COLS = ["Proc_Code","textbox9","textbox13",
                      "Count of Inv_Num","Sum of Crg_Amt","Sum of Paid_Amt"]

def process_pcd(file_bytes: bytes):
    df = read_first_sheet(file_bytes)
    o1 = df.drop(columns=["textbox27","textbox28","textbox29","textbox30"], errors="ignore")
    o1 = o1[(o1["Balance"]==0) & (o1["Crg_Amt"]!=0)][OUTPUT1_COLS].reset_index(drop=True)
    mask = df["Proc_Code"].astype(str).str.startswith("992", na=False)
    o2 = (df[mask]
          .groupby(["Proc_Code","textbox9","textbox13"], as_index=False)
          .agg(**{"Count of Inv_Num":("Inv_Num","count"),
                  "Sum of Crg_Amt":("Crg_Amt","sum"),
                  "Sum of Paid_Amt":("Paid_Amt","sum")})
          .sort_values("Count of Inv_Num", ascending=False)
          .reset_index(drop=True))[OUTPUT2_FINAL_COLS]
    return o1, o2

PCD1_COL_SPEC = {
    "Inv_Num":("text",12),"Svc_Date":("date",13),"Proc_Code":("text",12),
    "Modifier":(None,10),"Pat_Name":(None,20),"Pat_Num":("text",12),
    "Pat_DOB":("date",13),"textbox13":(None,14),"textbox11":(None,14),
    "textbox9":(None,14),"Phy_Name":(None,20),"Crg_Amt":("dollar",14),
    "Adj_Amt":("dollar",14),"Paid_Amt":("dollar",14),"Balance":("dollar",14),
}
PCD2_COL_SPEC = {
    "Proc_Code":("text",12),"textbox9":(None,14),"textbox13":(None,14),
    "Count of Inv_Num":("integer",16),"Sum of Crg_Amt":("dollar",16),"Sum of Paid_Amt":("dollar",16),
}

# ──────────────────────────────────────────────────────────────────────────────
#  Template Filler
# ──────────────────────────────────────────────────────────────────────────────

VISIT_TYPE_MAP  = {"P":"Private","W":"Work Comp","E":"EPS","M":"Misc"}
ALLOWED_EM_CODES = {"99201","99202","99203","99204","99205","99211","99212",
                    "99213","99214","99215","99201DOT","99201NON"}
TEMPLATE_COL = {
    "Practice":1,"Practice Description":2,"Clinic":3,"Visit Type":4,
    "Patient Number":5,"Visit Log Number":6,"Visit Identifier":7,
    "Visit Category 1":8,"Rendering Provider Credentials":9,"Provider Credentials":10,
    "Patient Sex at Birth":11,"Chart E/M Code":12,"Visit E/M Type":13,
    "Visit Category 2":14,"Visit Service Date":15,
    "Month, Year of Visit Service Date":16,"Primary Financial Class":17,
    "Patient Name":18,"Visit Arrival Status":19,"Charge is Rebill":20,
    "Visit Count":21,"Charge Amount":22,"Charge Proceture Code Quantity":23,
}

def _make_uid(pat_num, svc_date):
    if pd.isna(pat_num) or pd.isna(svc_date): return None
    pat = str(int(pat_num)) if isinstance(pat_num,float) else str(pat_num).strip()
    dt  = pd.to_datetime(svc_date, errors="coerce")
    return None if pd.isna(dt) else f"{pat}{dt.strftime('%m%d%Y')}"

def _creds(v):
    if pd.isna(v) or not str(v).strip(): return None
    tokens = str(v).split(",")[0].strip().split()
    return tokens[-1] if len(tokens)>=2 else None

def _em_type(code):
    if code in (None,"","-"): return "-"
    c = str(code).strip()
    if c in ("99201DOT","99201NON"): return "NON-E/M"
    if c.startswith("9921"): return "Established Patient"
    if c.startswith("9920"): return "New Patient"
    return "-"

def _filter_em(code):
    if pd.isna(code): return "-"
    c = str(code).strip()
    if c.endswith(".0"): c = c[:-2]
    return c if c in ALLOWED_EM_CODES else "-"

def _fmt_my(d):
    if pd.isna(d): return None
    dt = pd.to_datetime(d, errors="coerce")
    return dt.strftime("%b-%y") if not pd.isna(dt) else None

def _fmt_date(d):
    if pd.isna(d): return None
    dt = pd.to_datetime(d, errors="coerce")
    return dt.strftime("%m/%d/%Y") if not pd.isna(dt) else None


def build_template_sheet(file_map: dict, template_path: str):
    """file_map must contain: CNT_27, CNT_19, PAT_2, PCD_20, FIN_18"""
    cnt27 = read_first_sheet(file_map["CNT_27"])
    cnt19 = read_first_sheet(file_map["CNT_19"])
    pat2  = read_first_sheet(file_map["PAT_2"])
    pcd20 = read_first_sheet(file_map["PCD_20"])
    fin18 = read_first_sheet(file_map["FIN_18"])

    rebill_set = set(pd.to_numeric(fin18["New_Inv_Num"],errors="coerce").dropna().astype(int))
    cnt27["_uid"] = cnt27.apply(lambda r: _make_uid(r["Pat_Num"],r["Svc_Date"]),axis=1)
    cnt19["_uid"] = cnt19.apply(lambda r: _make_uid(r["Pat_Num"],r["Svc_Date"]),axis=1)
    cat_map   = dict(zip(cnt19["_uid"],         cnt19["Category"]))
    sex_map   = dict(zip(pat2["Patient_Number"], pat2["Sex"]))
    em_map    = dict(zip(pcd20["Inv_Num"],       pcd20["Proc_Code"]))
    qty_map   = pcd20.groupby("Inv_Num")["Proc_Code"].count().to_dict()

    wb = load_workbook(template_path)
    ws = wb.active
    C  = TEMPLATE_COL

    for i, row in cnt27.iterrows():
        r = i + 2
        pn, ln, inv = row.get("Pat_Num"), row.get("Log_Num"), row.get("Inv_Num")
        vt, rend    = row.get("Visit_Type"), row.get("Rendering_Phy")
        em          = _filter_em(em_map.get(inv)) if pd.notna(inv) else "-"
        cat         = cat_map.get(row["_uid"])

        ws.cell(r,C["Clinic"],        value=row.get("Clinic"))
        ws.cell(r,C["Visit Type"],    value=VISIT_TYPE_MAP.get(vt,vt))
        ws.cell(r,C["Patient Number"],   value=int(pn)  if pd.notna(pn)  else None)
        ws.cell(r,C["Visit Log Number"], value=int(ln)  if pd.notna(ln)  else None)
        ws.cell(r,C["Visit Identifier"], value=int(inv) if pd.notna(inv) else None)
        ws.cell(r,C["Visit Category 1"], value=cat)
        ws.cell(r,C["Visit Category 2"], value=cat)
        ws.cell(r,C["Rendering Provider Credentials"], value=rend)
        ws.cell(r,C["Provider Credentials"],           value=_creds(rend))
        ws.cell(r,C["Patient Sex at Birth"],            value=sex_map.get(pn))
        ws.cell(r,C["Chart E/M Code"],  value=em)
        ws.cell(r,C["Visit E/M Type"],  value=_em_type(em))
        ws.cell(r,C["Visit Service Date"],                value=_fmt_date(row.get("Svc_Date")))
        ws.cell(r,C["Month, Year of Visit Service Date"], value=_fmt_my(row.get("Svc_Date")))
        ws.cell(r,C["Primary Financial Class"], value=row.get("Class"))
        ws.cell(r,C["Patient Name"],            value=row.get("Pat_Name"))
        ws.cell(r,C["Visit Arrival Status"],    value=row.get("ArrivalStatus"))
        ws.cell(r,C["Charge is Rebill"],  value="Y" if (pd.notna(inv) and int(inv) in rebill_set) else "N")
        ws.cell(r,C["Visit Count"],       value=1)
        ws.cell(r,C["Charge Amount"],     value=row.get("Total_Charge"))
        ws.cell(r,C["Charge Proceture Code Quantity"], value=qty_map.get(inv) if pd.notna(inv) else None)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue(), len(cnt27)


# ──────────────────────────────────────────────────────────────────────────────
#  xlsxwriter helpers
# ──────────────────────────────────────────────────────────────────────────────

HDR_BG="  #1F3864"; STRIPE="#F2F2F2"; TOTAL_BG="#D9E1F2"

def build_formats(wb):
    base={"font_name":"Arial","font_size":10,"border":0}
    hdr=wb.add_format({**base,"bold":True,"font_color":"#FFFFFF","bg_color":"#1F3864",
                        "align":"center","valign":"vcenter","text_wrap":True})
    fmts={"header":hdr}
    for bt,bg,bold in [("white","#FFFFFF",False),("stripe","#F2F2F2",False),("total","#D9E1F2",True)]:
        def add(tag,nf=None,al="left",_bg=bg,_bo=bold,_bt=bt):
            f={**base,"bg_color":_bg,"align":al,"valign":"vcenter"}
            if _bo: f["bold"]=True
            if nf:  f["num_format"]=nf
            fmts[f"{_bt}_{tag}"]=wb.add_format(f)
        add("plain"); add("center",al="center")
        add("dollar",FMT_DOLLAR,"right"); add("integer",FMT_INTEGER,"center")
        add("date",FMT_DATE,"center");    add("text",FMT_TEXT,"left")
    return fmts

def pick_fmt(fmts,bt,ct):
    return fmts.get(f"{bt}_{ct}",fmts[f"{bt}_plain"])

def write_sheet(wb,fmts,name,df,spec,freeze="A2",is_arc=False):
    ws=wb.add_worksheet(name); cols=list(df.columns)
    ws.set_row(0,24)
    for ci,col in enumerate(cols):
        ws.write(0,ci,col,fmts["header"]); ws.set_column(ci,ci,spec.get(col,(None,14))[1])
    for ri,row in enumerate(df.itertuples(index=False)):
        er=ri+1; rv=list(row)
        bt="total" if (is_arc and rv[0]=="Total") else ("stripe" if ri%2 else "white")
        for ci,col in enumerate(cols):
            val=rv[ci]; ct=spec.get(col,(None,14))[0]; fmt=pick_fmt(fmts,bt,ct)
            if val is None or val=="": ws.write_blank(er,ci,None,fmt); continue
            try:    nan=pd.isna(val)
            except: nan=False
            if nan: ws.write_blank(er,ci,None,fmt); continue
            if   ct=="integer": ws.write_number(er,ci,int(round(float(val))),fmt)
            elif ct=="dollar":  ws.write_number(er,ci,float(val),fmt)
            elif ct=="date":
                if hasattr(val,"year"): ws.write_datetime(er,ci,val,fmt)
                else:                   ws.write(er,ci,str(val),fmts[f"{bt}_plain"])
            elif ct=="text": ws.write_string(er,ci,str(val),fmt)
            else:
                if isinstance(val,(int,float)): ws.write_number(er,ci,val,fmt)
                else:                           ws.write(er,ci,str(val),fmt)
    ws.freeze_panes(freeze); ws.autofilter(0,0,len(df),len(cols)-1); ws.set_zoom(90)


# ──────────────────────────────────────────────────────────────────────────────
#  Main pipeline
# ──────────────────────────────────────────────────────────────────────────────

def run_combined(file_map: dict):
    import xlsxwriter
    from openpyxl import load_workbook as _lw

    errors, sheets_written = [], []
    arc_df = pcd_o1 = pcd_o2 = tmpl_bytes = None

    # ARC_18
    if "ARC_18" in file_map:
        try:    arc_df = compute_arc_summary(file_map["ARC_18"], cnt27_bytes=file_map.get("CNT_27"))
        except Exception as e: errors.append(f"ARC_18 error: {e}")
    else:
        errors.append("ARC_18 not uploaded — Begin & Ending AR sheet skipped.")

    # PCD_20
    if "PCD_20" in file_map:
        try:    pcd_o1, pcd_o2 = process_pcd(file_map["PCD_20"])
        except Exception as e: errors.append(f"PCD_20 error: {e}")
    else:
        errors.append("PCD_20 not uploaded — Closed Report & E&M sheets skipped.")

    # Template filler
    tmpl_path   = find_template()
    tmpl_needed = {"CNT_27","CNT_19","PAT_2","PCD_20","FIN_18"}
    missing     = tmpl_needed - set(file_map.keys())
    if not tmpl_path:
        errors.append("No template file in ./template/ folder — Charge Detail Report skipped.")
    elif missing:
        errors.append(f"Charge Detail Report skipped — missing: {', '.join(sorted(missing))}")
    else:
        try:    tmpl_bytes, _ = build_template_sheet(file_map, tmpl_path)
        except Exception as e: errors.append(f"Template filler error: {e}")

    if arc_df is None and pcd_o1 is None and tmpl_bytes is None:
        raise ValueError("No data could be processed.\n" + "\n".join(errors))

    # Build entirely in memory — no file written to disk
    buf = io.BytesIO()
    with xlsxwriter.Workbook(buf,{"constant_memory":False,"strings_to_numbers":False,
                                   "default_date_format":"mm/dd/yyyy","in_memory":True}) as wb:
        fmts = build_formats(wb)
        if arc_df is not None:
            write_sheet(wb,fmts,"Begin & Ending AR by Clinic",arc_df,ARC_COL_SPEC,"D2",is_arc=True)
            sheets_written.append("Begin & Ending AR by Clinic")
        if pcd_o1 is not None:
            write_sheet(wb,fmts,"SJC_Closed Report",     pcd_o1,PCD1_COL_SPEC,"A2")
            write_sheet(wb,fmts,"SJC_E&M_Visit Summary", pcd_o2,PCD2_COL_SPEC,"A2")
            sheets_written += ["SJC_Closed Report","SJC_E&M_Visit Summary"]
    buf.seek(0)

    if tmpl_bytes:
        try:
            master  = _lw(buf)
            tsrc    = _lw(io.BytesIO(tmpl_bytes)).active
            new_ws  = master.create_sheet(title="Charge Detail Report")
            for row in tsrc.iter_rows():
                for cell in row:
                    nc = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        nc.font=cell.font.copy(); nc.border=cell.border.copy()
                        nc.fill=cell.fill.copy(); nc.number_format=cell.number_format
                        nc.alignment=cell.alignment.copy()
            for cl,dim in tsrc.column_dimensions.items():
                new_ws.column_dimensions[cl].width = dim.width
            buf = io.BytesIO(); master.save(buf); buf.seek(0)
            sheets_written.append("Charge Detail Report")
        except Exception as e:
            errors.append(f"Could not append Charge Detail Report: {e}")

    return buf, sheets_written, errors


# ──────────────────────────────────────────────────────────────────────────────
#  Flask routes
# ──────────────────────────────────────────────────────────────────────────────

SESSION_STORE: dict = {}   # session_id → raw bytes held in memory


@app.route("/")
def index():
    tmpl = find_template()
    return render_template("index.html", template_name=os.path.basename(tmpl) if tmpl else None)


@app.route("/process", methods=["POST"])
def process():
    uploaded = request.files.getlist("raw_files")
    if not uploaded:
        return jsonify({"error": "No files uploaded."}), 400

    file_map, unmatched = {}, []
    for f in uploaded:
        if not f.filename: continue
        key = resolve_file_key(f.filename)
        if key:
            file_map[key] = f.read()
        else:
            unmatched.append(f.filename)

    if not file_map:
        return jsonify({"error": "None of the uploaded files matched a known data source (ARC_18, PCD_20, CNT_27, CNT_19, PAT_2, FIN_18)."}), 400

    session_id = str(uuid.uuid4())[:8]
    try:
        buf, sheets, warnings = run_combined(file_map)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    SESSION_STORE[session_id] = buf.read()   # store raw bytes, nothing written to disk

    if unmatched:
        warnings.append(f"Unrecognised files (skipped): {', '.join(unmatched)}")

    return jsonify({
        "session_id": session_id,
        "sheets":     sheets,
        "warnings":   warnings,
        "matched":    list(file_map.keys()),
        "filename":   f"combined_output_{session_id}.xlsx",
    })


@app.route("/download/<session_id>")
def download(session_id):
    session_id = re.sub(r"[^a-zA-Z0-9_-]","",session_id)
    data = SESSION_STORE.pop(session_id, None)   # consume once then discard
    if data is None:
        return "File not found or already downloaded", 404
    return send_file(io.BytesIO(data), as_attachment=True,
                     download_name="SJC_Combined_Report.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5050, debug=False, threaded=True)
